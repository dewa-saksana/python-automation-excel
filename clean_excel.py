# clean_excel.py
import openpyxl

wb = openpyxl.load_workbook(data.xlsx)
ws = nwb.active

# Delete empty line
for row in range(ws.max_row, 0, -1):
  if all([cell.value is None for cell in ws[row]]):
    ws.delete_rows(row)

wb.save("data_cleaned.xlsx")
print("Cleaning complete."}
  
